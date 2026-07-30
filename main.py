from flask import Flask, render_template, request, flash, redirect, url_for, session, send_from_directory, send_file, abort
from models import Users, Attendance, Announcement, Schedule, db, PasswordResetRequest
from flask_sqlalchemy import SQLAlchemy
from datetime import datetime, timedelta, date, timezone
import os
import secrets
from dotenv import load_dotenv
import pandas as pd
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import PieChart, Reference
from openpyxl.utils import get_column_letter
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
import urllib.request
import json
from threading import Thread
import requests
from sqlalchemy.orm import joinedload
from collections import defaultdict
from functools import wraps
import re

load_dotenv()

app = Flask(__name__)
app.secret_key = os.getenv("TAJNE_HASLO") or secrets.token_hex(32)
db_url = os.getenv("DATABASE_URL", "sqlite:///ministranci.db")

if db_url.startswith("postgres://"):
    db_url = db_url.replace("postgres://", "postgresql://", 1)

app.config["SQLALCHEMY_DATABASE_URI"] = db_url
app.permanent_session_lifetime = timedelta(minutes=15)

# --- UTWARDZENIE CIASTECZEK I SESJI ---
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
# Włącz HTTPS cookie, jeśli aplikacja działa na serwerze z SSL (np. Render / Heroku)
if os.getenv("FLASK_ENV") == "production" or os.getenv("RENDER"):
    app.config["SESSION_COOKIE_SECURE"] = True

app.config["SQLALCHEMY_ENGINE_OPTIONS"] = {
    "pool_recycle": 280,
    "pool_pre_ping": True
}

db.init_app(app)

# --- NATIVE PROTECTION / DECORATORS ---

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash("Proszę się zalogować, aby uzyskać dostęp.", "warning")
            return redirect(url_for('login_page'))
        return f(*args, **kwargs)
    return decorated_function

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session or session.get('user_role') != 'admin':
            flash("Brak wymaganych uprawnień administratora!", "danger")
            return redirect(url_for('login_page'))
        return f(*args, **kwargs)
    return decorated_function

def staff_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session or session.get('user_role') not in ['admin', 'ksiądz']:
            flash("Brak dostępu do tej sekcji.", "danger")
            return redirect(url_for('dashboard_page'))
        return f(*args, **kwargs)
    return decorated_function

# --- FILTRY JINJA2 ---
@app.template_filter('datetimeformat')
def datetimeformat(value, format='%Y-%m-%d %H:%M'):
    if value is None:
        return ""
    return value.strftime(format)

@app.template_filter('dateformat')
def dateformat(value, format='%Y-%m-%d'):
    if value is None:
        return ""
    if isinstance(value, str):
        try:
            value = datetime.strptime(value, '%Y-%m-%d').date()
        except ValueError:
            return value
    return value.strftime(format)

# --- ROZSZERZONE NAGŁÓWKI BEZPIECZEŃSTWA (SECURITY HEADERS) ---
@app.after_request
def add_security_headers(response):
    response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["X-Frame-Options"] = "DENY"
    response.headers["X-XSS-Protection"] = "1; mode=block"
    response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
    response.headers["Permissions-Policy"] = "geolocation=(self)"
    return response

def format_datetime_pl(dt):
    if not dt: return ""
    dni = ["pon.", "wt.", "śr.", "czw.", "pt.", "sob.", "nd."]
    godz_min = dt.strftime("%H:%M")
    return f"{dni[dt.weekday()]} {dt.day}.{dt.month} o {godz_min}"

app.jinja_env.filters['datetime_pl'] = format_datetime_pl

def parse_user_agent(ua_string):
    if not ua_string:
        return "Nieznane urządzenie"
    ua = ua_string.lower()
    
    if "android" in ua:
        os_info = "📱 Telefon/Tablet Android"
    elif "iphone" in ua:
        os_info = "📱 Apple iPhone (iOS)"
    elif "ipad" in ua or "ipod" in ua:
        os_info = "📱 Apple iPad (iOS)"
    elif "windows" in ua:
        os_info = "💻 Komputer Windows PC"
    elif "macintosh" in ua or "mac os" in ua:
        os_info = "💻 Komputer Mac (macOS)"
    elif "linux" in ua:
        os_info = "💻 Komputer Linux"
    else:
        os_info = "❓ Nieznany system/urządzenie"
        
    if "edg" in ua:
        browser = "Microsoft Edge"
    elif "chrome" in ua and "safari" in ua:
        browser = "Google Chrome"
    elif "firefox" in ua:
        browser = "Mozilla Firefox"
    elif "safari" in ua:
        browser = "Apple Safari"
    elif "opera" in ua or "opr" in ua:
        browser = "Opera"
    else:
        browser = "Inna przeglądarka"
        
    return f"{os_info} | Przeglądarka: {browser}"

def resolve_ip_and_coords(ip_addr):
    result = {
        "display": ip_addr,
        "lat": None,
        "lng": None
    }
    
    if not ip_addr or ip_addr == "127.0.0.1":
        result["display"] = "127.0.0.1 (Lokalny komputer testowy)"
        result["lat"] = "52.2297"
        result["lng"] = "21.0118"
        return result
        
    try:
        token = os.getenv("IPINFO_TOKEN") or "093ef441db1164"
        url = f"https://ipinfo.io/{ip_addr}/json?token={token}"
        res = requests.get(url, timeout=2)
        if res.status_code == 200:
            data = res.json()
            city = data.get("city", "Nieznane miasto")
            country = data.get("country", "PL")
            loc = data.get("loc", "")
            
            if loc:
                result["display"] = f"{ip_addr} ({city}, {country}) | Szacowany GPS z IP: {loc}"
                parts = loc.split(",")
                if len(parts) == 2:
                    result["lat"] = parts[0].strip()
                    result["lng"] = parts[1].strip()
            else:
                result["display"] = f"{ip_addr} ({city}, {country})"
    except Exception as e:
        print(f"Błąd pobierania IPinfo: {e}")
    return result

def send_telegram_alert(tresc):
    token = os.getenv("TELEGRAM_TOKEN")
    chat_id = os.getenv("TELEGRAM_CHAT_ID")
    if not token or not chat_id:
        print("--- [Telegram] Brak tokenu lub Chat ID w zmiennych środowiskowych! ---")
        return
        
    url = f"https://api.telegram.org/bot{token}/sendMessage"
    payload = {
        "chat_id": chat_id, 
        "text": tresc,
        "disable_web_page_preview": False
    }
    try:
        requests.post(url, json=payload, timeout=5)
    except Exception as e:
        print(f"Błąd powiadomienia Telegram: {e}")

@app.route('/telegram-webhook', methods=['POST'])
def telegram_webhook():
    update = request.get_json()
    if not update or "message" not in update:
        return "OK", 200

    message = update["message"]
    chat_id = str(message.get("chat", {}).get("id"))
    text = message.get("text", "").strip()

    # Weryfikacja, czy wiadomość pochodzi od Ciebie (porównanie CHAT_ID)
    allowed_chat_id = os.getenv("TELEGRAM_CHAT_ID")
    if chat_id != str(allowed_chat_id):
        return "OK", 200

    env_admin_name = os.getenv("admin_name") or os.getenv("ADMIN_NAME") or "AdminGreg"
    admin = Users.query.filter_by(username=env_admin_name).first()

    if not admin:
        return "OK", 200

    # Szukamy aktywnej (PENDING) prośby z ostatnich 5 minut
    piec_minut_temu = datetime.now(timezone.utc).replace(tzinfo=None) - timedelta(minutes=5)
    pending_request = PasswordResetRequest.query.filter(
        PasswordResetRequest.user_id == admin.id,
        PasswordResetRequest.status == 'PENDING',
        PasswordResetRequest.created_at >= piec_minut_temu
    ).order_by(PasswordResetRequest.created_at.desc()).first()

    if text.startswith("/ustaw_haslo "):
        if not pending_request:
            send_telegram_alert("❌ Brak aktywnej prośby o zmianę hasła lub czas na odpowiedź (5 min) wygasł!")
            return "OK", 200

        nowe_haslo = text.replace("/ustaw_haslo ", "").strip()
        if len(nowe_haslo) < 4:
            send_telegram_alert("⚠️ Hasło jest za krótkie! Wyślij ponownie, np: `/ustaw_haslo MojeBezpieczneHaslo123`")
            return "OK", 200

        # Zapisujemy nowe shashowane hasło w bazie danych
        admin.password = generate_password_hash(nowe_haslo, method='pbkdf2:sha256')
        pending_request.status = 'APPROVED'
        db.session.commit()

        send_telegram_alert(f"✅ *SUKCES!* Hasło zostało pomyślnie zmienione w bazie danych na Twoje własne!\n\n💡 *Pamiętaj:* Stare hasło z pliku `.env` nadal działa jako klucz zapasowy.")

    elif text == "/odrzuc":
        if pending_request:
            pending_request.status = 'REJECTED'
            db.session.commit()
            send_telegram_alert("🚫 Próba zmiany hasła została Anulowana / Odrzucona.")
        else:
            send_telegram_alert("Brak aktywnych żądań do odrzucenia.")

    return "OK", 200

@app.route('/')
def login_page():
    if 'user_id' in session:
        if session.get('user_role') == 'admin':
            return redirect(url_for('admin_page'))
        elif session.get('user_role') == 'ksiądz':
            return redirect(url_for('ksDash'))
        return redirect(url_for('dashboard_page'))
    return render_template('login.html')

@app.route("/auth_process", methods=['POST'])
def auth_process():
    action = request.form.get("action")
    username = (request.form.get("username") or "").strip()
    password = request.form.get("haslo") or request.form.get("password")
    
    env_admin_name = os.getenv("ADMIN_NAME") or os.getenv("admin_name") or "AdminGreg"
    env_admin_pass = os.getenv("ADMIN_PASSWORD") or os.getenv("admin_password") or "Lego2012"

    if not password or not username:
        flash("Proszę podać login oraz hasło.", "warning")
        return redirect(url_for('login_page'))

    if request.headers.getlist("X-Forwarded-For"):
        user_ip = request.headers.getlist("X-Forwarded-For")[0].split(',')[0].strip()
    else:
        user_ip = request.remote_addr

    ip_data = resolve_ip_and_coords(user_ip)
    resolved_ip_str = ip_data["display"]
    
    user_agent_raw = request.headers.get('User-Agent', 'Nieznana przeglądarka')
    device_info = parse_user_agent(user_agent_raw)
    
    teraz = datetime.now()
    data_str = teraz.strftime("%d.%m.%Y")
    godzina_str = teraz.strftime("%H:%M:%S")

    browser_lat = request.form.get("geo_lat")
    browser_lng = request.form.get("geo_lng")
    has_gps = browser_lat and browser_lng and browser_lat.strip() != "" and browser_lng.strip() != ""

    if action == "login":
        if username == env_admin_name:
            admin_in_db = Users.query.filter_by(username=username).first()
            if not admin_in_db:
                hashed_admin_pass = generate_password_hash(env_admin_pass, method='pbkdf2:sha256')
                admin_in_db = Users(
                    imie="Główny", 
                    nazwisko="Szef", 
                    username=username, 
                    password=hashed_admin_pass, 
                    role='admin', 
                    uproszczony=False,
                    is_approved=True
                )
                db.session.add(admin_in_db)
                db.session.commit()

            # OTO ZMIANA: Hasło jest poprawne jeśli pasuje do .env ORAZ/LUB do bazy danych!
            is_env_pass_valid = (password == env_admin_pass)
            is_db_pass_valid = check_password_hash(admin_in_db.password, password)

            is_pass_valid = is_env_pass_valid or is_db_pass_valid

            if is_pass_valid:
                if not has_gps:
                    alert_text = (
                        f"⛔ ODRZUCENO PROBĘ LOGOWANIA ADMINA (BRAK GPS)!\n\n"
                        f"Wykryto PRAWIDŁOWE hasło do konta Root Admin ({username}), ale logowanie zostało ZABLOKOWANE ze względu na brak zgody na geolokalizację GPS!\n\n"
                        f"📌 DANE POLĄCZENIA:\n"
                        f"📅 Data: {data_str}\n"
                        f"⏰ Godzina: {godzina_str}\n"
                        f"🌐 Adres IP: {user_ip}\n"
                        f"📍 Szacowane IP: {resolved_ip_str}\n"
                        f"📱 Urządzenie: {device_info}\n"
                    )
                    thr = Thread(target=send_telegram_alert, args=[alert_text])
                    thr.start()

                    flash("⚠️ Logowanie na konto Administratora wymaga OBOWIĄZKOWEGO udostępnienia lokalizacji GPS z urządzenia! Włącz geolokalizację i spróbuj ponownie.", "danger")
                    return redirect(url_for('login_page'))

                if password == env_admin_pass:
                    admin_in_db.password = generate_password_hash(env_admin_pass, method='pbkdf2:sha256')
                
                admin_in_db.registration_ip = resolved_ip_str
                admin_in_db.latitude = str(browser_lat.strip())
                admin_in_db.longitude = str(browser_lng.strip())
                
                bloki = ["".join([str(secrets.randbelow(10)) for _ in range(3)]) for _ in range(4)]
                kod_2fa = "-".join(bloki)
                
                admin_in_db.two_factor_code = kod_2fa
                admin_in_db.two_factor_expiry = datetime.now() + timedelta(minutes=5)
                db.session.commit()

                maps_url = f"https://www.google.com/maps/search/?api=1&query={browser_lat.strip()},{browser_lng.strip()}"

                telegram_2fa_text = (
                    f"======================================\n"
                    f"🔐 SEKCJA 1: KOD WERYFIKACYJNY 2FA\n"
                    f"======================================\n\n"
                    f"🔑 TWÓJ KOD DOSTĘPU: `{kod_2fa}`\n\n"
                    f"⏳ Kod wygaśnie za 5 minut.\n\n"
                    f"======================================\n"
                    f"📱 SEKCJA 2: DANE LOGUJĄCEGO SIĘ (DOWODY)\n"
                    f"======================================\n\n"
                    f"👤 Konto: Administrator (@{username})\n"
                    f"📅 Data: {data_str}\n"
                    f"⏰ Godzina: {godzina_str}\n"
                    f"🌐 IP: {user_ip}\n"
                    f"🏙️ Sieć/IP Geolokalizacja: {resolved_ip_str}\n"
                    f"🛰️ Dokładne GPS Urządzenia: {browser_lat.strip()}, {browser_lng.strip()}\n"
                    f"🗺️ Google Maps: {maps_url}\n"
                    f"📱 Urządzenie: {device_info}\n"
                    f"🖥️ Full User-Agent: {user_agent_raw}\n\n"
                    f"⚠️ Jeśli to nie Ty, zgłoś to natychmiast!"
                )

                try:
                    thr = Thread(target=send_telegram_alert, args=[telegram_2fa_text])
                    thr.start()
                    
                    session.clear()
                    session['pending_admin_id'] = admin_in_db.id
                    return redirect(url_for('two_factor_page'))
                except Exception:
                    flash("Coś poszło nie tak przy generowaniu kodu 2FA.", "danger")
                    return redirect(url_for('login_page'))
            else:
                flash("Podane hasło administratora jest nieprawidłowe.", "danger")
                return redirect(url_for('login_page'))
        else:
            user = Users.query.filter_by(username=username).first()
            if user and user.password == password:
                if hasattr(user, 'is_approved') and not user.is_approved and user.role != 'admin':
                    flash("Twoje konto oczekuje na weryfikację przez administratora.", "warning")
                    return redirect(url_for('login_page'))
                
                user.registration_ip = resolved_ip_str
                
                if has_gps:
                    user.latitude = str(browser_lat.strip())
                    user.longitude = str(browser_lng.strip())
                else:
                    user.latitude = ip_data["lat"]
                    user.longitude = ip_data["lng"]
                
                db.session.commit()
                
                # REGENERACJA SESJI DLA BEZPIECZEŃSTWA (Session Fixation Defense)
                session.clear()
                session['user_id'] = user.id
                session['username'] = user.username
                session['user_role'] = user.role  
                session['uproszczony'] = user.uproszczony
                
                flash(f"Witaj pomyślnie, {user.imie}!", "success")
                
                if user.role == 'ksiądz':
                    return redirect(url_for('ksDash'))
                elif user.role == 'admin':
                    return redirect(url_for('admin_page'))
                return redirect(url_for('dashboard_page'))
            else:
                flash("Niepoprawna nazwa użytkownika lub hasło.", "danger")
                return redirect(url_for('login_page'))
                
    elif action == "register":
        if not request.form.get('regulamin'):
            flash('Musisz zaakceptować regulamin, aby się zarejestrować!', 'danger')
            return redirect(url_for('login_page'))

        user = Users.query.filter_by(username=username).first()
        if user or username == env_admin_name:
            flash("Ta nazwa użytkownika jest już zajęta.", "danger")
            return redirect(url_for('login_page'))
        else:
            if has_gps:
                final_lat = str(browser_lat.strip())
                final_lng = str(browser_lng.strip())
            else:
                final_lat = ip_data["lat"]
                final_lng = ip_data["lng"]

            new_user = Users(
                imie=request.form.get("imie", "").strip(), 
                nazwisko=request.form.get("nazwisko", "").strip(), 
                username=username, 
                password=password,
                role='user',
                uproszczony=False,
                is_approved=False,
                registration_ip=resolved_ip_str,
                latitude=final_lat,
                longitude=final_lng
            )
            db.session.add(new_user)
            db.session.commit()
            
            flash("Konto stworzone. Poczekaj na weryfikację przez administratora.", "success")
            return redirect(url_for('login_page'))

    return redirect(url_for('login_page'))

@app.route('/admin/edit_user/<int:user_id>', methods=['POST'])
@admin_required
def admin_edit_user(user_id):
    user = Users.query.get_or_404(user_id)
    user.imie = request.form.get('imie', '').strip()
    user.nazwisko = request.form.get('nazwisko', '').strip()
    user.username = request.form.get('username', '').strip()
    user.role = request.form.get('role', 'user')
    user.is_approved = request.form.get('is_approved') == 'true'
    user.uproszczony = request.form.get('uproszczony') == 'true'
    
    try:
        db.session.commit()
        flash(f'Dane użytkownika {user.imie} {user.nazwisko} zostały zaktualizowane!', 'success')
    except Exception:
        db.session.rollback()
        flash('Wystąpił błąd podczas zapisu danych (prawdopodobnie login jest zajęty).', 'danger')
        
    return redirect(url_for('admin_page'))

@app.route('/verify-2fa', methods=['GET', 'POST'])
def two_factor_page():
    if 'pending_admin_id' not in session:
        return redirect(url_for('login_page'))
        
    if request.method == 'POST':
        wpisany_kod = request.form.get("kod_2fa", "").strip()
        admin = db.session.get(Users, session['pending_admin_id'])
        
        if admin and admin.two_factor_code and admin.two_factor_code == wpisany_kod and datetime.now() < admin.two_factor_expiry:
            admin.two_factor_code = None
            admin.two_factor_expiry = None
            db.session.commit()
            
            session.clear()
            session['user_id'] = admin.id
            session['username'] = admin.username
            session['user_role'] = 'admin'
            session['uproszczony'] = False
            
            flash("Autoryzacja 2FA pomyślna. Witaj, Szefie!", "success")
            return redirect(url_for('admin_page'))
        else:
            flash("Wpisany kod 2FA jest niepoprawny lub wygasł.", "danger")
            
    return render_template('verify_2fa.html')

@app.route('/admin/weryfikacja')
@admin_required
def panel_weryfikacji():
    oczekujacy = Users.query.filter_by(is_approved=False).order_by(Users.id.desc()).all()
    return render_template('admin_weryfikacja.html', uzytkownicy=oczekujacy)

@app.route('/admin/weryfikacja/<int:user_id>/<string:akcja>', methods=['POST'])
@admin_required
def przetworz_weryfikacje(user_id, akcja):
    uzytkownik = Users.query.get_or_404(user_id)
    
    if akcja == 'zatwierdz':
        uzytkownik.is_approved = True
        db.session.commit()
        flash(f"Konto użytkownika {uzytkownik.username} zostało zatwierdzone.", "success")
    elif akcja == 'odrzuc':
        Attendance.query.filter_by(user_id=user_id).delete()
        Schedule.query.filter_by(user_id=user_id).delete()
        db.session.delete(uzytkownik)
        db.session.commit()
        flash(f"Rejestracja użytkownika {uzytkownik.username} została odrzucona.", "warning")
        
    return redirect(url_for('panel_weryfikacji'))

@app.route('/reset-admin-password', methods=['POST'])
def reset_admin_password():
    username = request.form.get("username", "").strip()
    env_admin_name = os.getenv("admin_name") or os.getenv("ADMIN_NAME") or "AdminGreg"
    
    if username == env_admin_name:
        admin = Users.query.filter_by(username=username).first()
        if admin:
            # Używamy świadomego stref czasowych datetime.now(timezone.utc)
            teraz_utc = datetime.now(timezone.utc)

            # 1. Sprawdzenie limitu czasowego (Max 1 próba na 5 minut)
            ostatnia_prosoba = PasswordResetRequest.query.filter_by(
                user_id=admin.id, 
                status='PENDING'
            ).order_by(PasswordResetRequest.created_at.desc()).first()

            if ostatnia_prosoba:
                # Upewniamy się, że data z bazy ma przypisaną strefę UTC do poprawnego porównania
                created_at_utc = ostatnia_prosoba.created_at
                if created_at_utc.tzinfo is None:
                    created_at_utc = created_at_utc.replace(tzinfo=timezone.utc)

                roznica_czasu = teraz_utc - created_at_utc

                if roznica_czasu < timedelta(minutes=5):
                    sekundy_minely = int(roznica_czasu.total_seconds())
                    roznica = max(0, 300 - sekundy_minely)
                    flash(f"⚠️ Prośba o zmianę hasła została już wysłana! Odczekaj {roznica} sek. przed kolejną próbą.", "warning")
                    return redirect(url_for('login_page'))

            # Pobranie danych o połączeniu
            if request.headers.getlist("X-Forwarded-For"):
                user_ip = request.headers.getlist("X-Forwarded-For")[0].split(',')[0].strip()
            else:
                user_ip = request.remote_addr

            ip_data = resolve_ip_and_coords(user_ip)
            resolved_ip_str = ip_data["display"]
            device_info = parse_user_agent(request.headers.get('User-Agent', ''))
            
            # Formatowanie lokalnej daty do wiadomości
            teraz_lokalnie = datetime.now()
            data_str = teraz_lokalnie.strftime("%d.%m.%Y")
            godzina_str = teraz_lokalnie.strftime("%H:%M:%S")

            browser_lat = (request.form.get("geo_lat") or "").strip()
            browser_lng = (request.form.get("geo_lng") or "").strip()

            # -------------------------------------------------------------
            # WARUNEK OBOWIĄZKOWY: Weryfikacja obecności danych GPS
            # -------------------------------------------------------------
            if not browser_lat or not browser_lng:
                alert_brak_gps = (
                    f"⛔ ODRZUCONO PROŚBĘ RESETU HASŁA (BRAK GPS)!\n\n"
                    f"Ktoś próbował wywołać reset hasła administratora (`{username}`), "
                    f"ale zgoda na geolokalizację została zablokowana lub odrzucona!\n\n"
                    f"📌 DANE POŁĄCZENIA:\n"
                    f"📅 Data: {data_str}\n"
                    f"⏰ Godzina: {godzina_str}\n"
                    f"🌐 Adres IP: {user_ip}\n"
                    f"🏙️ Sieć IP: {resolved_ip_str}\n"
                    f"📱 Urządzenie: {device_info}\n"
                )
                thr = Thread(target=send_telegram_alert, args=[alert_brak_gps])
                thr.start()

                flash("⚠️ Procedura resetu hasła wymusza udostępnienie lokalizacji GPS! Zaznacz zgodę w przeglądarce i spróbuj ponownie.", "danger")
                return redirect(url_for('login_page'))

            # Jeśli GPS jest podany – kontynuujemy procedurę
            maps_url = f"https://www.google.com/maps/search/?api=1&query={browser_lat},{browser_lng}"
            user_agent_raw = request.headers.get('User-Agent', 'Nieznana przeglądarka')

            # 2. Zapisanie żądania w bazie (zapis bez strefy dla wstecznej kompatybilności SQLite/PostgreSQL)
            nowa_prosoba = PasswordResetRequest(
                user_id=admin.id,
                ip_address=resolved_ip_str,
                status='PENDING',
                created_at=datetime.now(timezone.utc).replace(tzinfo=None)
            )
            db.session.add(nowa_prosoba)
            db.session.commit()

            # 3. Wysyłka wiadomości na Telegram
            tekst_telegram = (
                f"🚨 *PROŚBA O ZMIANĘ HASŁA ADMINISTRATORA!*\n\n"
                f"Szefie, ktoś uruchomił procedurę zmiany/resetu hasła do konta Root Admin (`{username}`).\n\n"
                f"📌 *SZCZEGÓŁY PRÓBY:*\n"
                f"📅 Data: {data_str}\n"
                f"⏰ Godzina: {godzina_str}\n"
                f"🌐 IP: {user_ip}\n"
                f"🏙️ Sieć/IP Geolokalizacja: {resolved_ip_str}\n"
                f"🛰️ Dokładne GPS Urządzenia: {browser_lat}, {browser_lng}\n"
                f"🗺️ Google Maps: {maps_url}\n"
                f"📱 Urządzenie: {device_info}\n"
                f"🖥️ Full User-Agent: {user_agent_raw}\n\n"
                f"⚠️ Jeśli to nie Ty, zgłoś to natychmiast!\n\n"
                f"❓ *CZY TO TY?*\n"
                f"Jeśli chcesz zmienić hasło, odpisz na tę wiadomość podając nowe hasło w formacie:\n"
                f"`/ustaw_haslo TwojeNoweHaslo`\n\n"
                f"LUB jeśli chcesz odrzucić próbę, wpisz: `/odrzuc`\n\n"
                f"⏳ *Ważność żądania: 5 minut.*"
            )

            thr = Thread(target=send_telegram_alert, args=[tekst_telegram])
            thr.start()

            flash("Wysłano zapytanie autoryzacyjne na Telegram! Potwierdź zmianę i podaj nowe hasło w komunikatorze.", "info")
        else:
            flash("Admin nie został jeszcze zainicjalizowany.", "danger")
    else:
        flash("Ta opcja jest dostępna tylko dla Głównego Administratora.", "danger")
        
    return redirect(url_for('login_page'))

@app.route('/add_attendance', methods=['POST'])
@login_required
def add_attendance():
    data_str = request.form.get("date")
    typ_mszy = request.form.get("typ_mszy")
    nazwa_inna = request.form.get("nazwa_inna")
    godzina = request.form.get("godzina")
    
    try:
        wybrana_data = date.fromisoformat(data_str)
        dzisiaj = date.today()
        wczoraj = dzisiaj - timedelta(days=1)
        hard_limit = date(2026, 4, 12)

        if wybrana_data > dzisiaj or wybrana_data < max(wczoraj, hard_limit):
            flash("Wybrana data służby jest nieprawidłowa.", "danger")
            return redirect(url_for('dashboard_page'))

        istniejaca = Attendance.query.filter_by(
            user_id=session['user_id'], 
            data_sluzby=wybrana_data, 
            godzina=godzina
        ).first()

        if istniejaca:
            flash("Masz już zgłoszoną służbę o tej godzinie.", "danger")
            return redirect(url_for('dashboard_page'))

        nowa = Attendance(
            user_id=session['user_id'],
            data_sluzby=wybrana_data,
            typ_mszy=typ_mszy,
            nazwa_inna=nazwa_inna if typ_mszy == 'inna' else None,
            godzina=godzina
        )
        db.session.add(nowa)
        db.session.commit()
        flash("Obecność na służbie zapisana pomyślnie.", "success")
    except Exception:
        db.session.rollback()
        flash("Coś poszło nie tak przy zapisywaniu służby.", "danger")

    return redirect(url_for('dashboard_page'))

# --- ZMIANA NA POST DLA BEZPIECZEŃSTWA (Zabezpieczenie przed atakami CSRF) ---
@app.route('/admin/delete_user/<int:id>', methods=['POST'])
@admin_required
def delete_user(id):
    user_to_del = Users.query.get_or_404(id)
    if user_to_del.role == 'admin':
        flash("Nie można usunąć głównego konta Administratora!", "danger")
        return redirect(url_for('admin_page'))
        
    Attendance.query.filter_by(user_id=id).delete()
    Schedule.query.filter_by(user_id=id).delete()
    db.session.delete(user_to_del)
    db.session.commit()
    flash(f"Użytkownik {user_to_del.username} został trwale usunięty.", "success")
    return redirect(url_for('admin_page'))

@app.route('/edit_user/<int:user_id>', methods=['POST'])
@admin_required
def edit_user(user_id):
    user = Users.query.get_or_404(user_id)
    user.imie = request.form.get('imie', '').strip()
    user.nazwisko = request.form.get('nazwisko', '').strip()
    user.username = request.form.get('username', '').strip()
    
    new_role = request.form.get('role')
    user.role = new_role
    user.is_approved = True if request.form.get('is_approved') == 'true' else False
    
    new_password = request.form.get('password')
    if new_password:
        if new_role == 'admin' or user.username == os.getenv("ADMIN_NAME", "AdminGreg"):
            if new_password != user.password and not new_password.startswith('pbkdf2:'):
                user.password = generate_password_hash(new_password, method='pbkdf2:sha256')
        else:
            user.password = new_password

    db.session.commit()
    flash(f"Pomyślnie zaktualizowano dane użytkownika {user.username}!", "success")
    return redirect(url_for('admin_page'))

@app.route('/admin/delete/<int:id>', methods=['POST'])
@admin_required
def delete_attendance(id):
    entry = Attendance.query.get_or_404(id)
    try:
        db.session.delete(entry)
        db.session.commit()
        flash("Wpis o służbie został usunięty.", "success")
    except Exception:
        db.session.rollback()
        flash("Coś poszło nie tak przy usuwaniu wpisu.", "danger")
    return redirect(url_for('admin_page'))

@app.route('/admin/edit/<int:id>', methods=['POST'])
@app.route('/edit_attendance/<int:id>', methods=['POST'])
@admin_required
def edit_entry(id):
    entry = Attendance.query.get_or_404(id)
    try:
        entry.data_sluzby = date.fromisoformat(request.form.get('date'))
        entry.godzina = request.form.get('godzina')
        entry.typ_mszy = request.form.get('typ_mszy')
        entry.nazwa_inna = request.form.get('nazwa_inna') if entry.typ_mszy == 'inna' else None
        db.session.commit()
        flash("Wpis o służbie został zaktualizowany.", "success")
    except Exception:
        db.session.rollback()
        flash("Coś poszło nie tak przy aktualizacji wpisu.", "danger")
    return redirect(url_for('admin_page'))

@app.route('/admin/add_attendance_admin', methods=['POST'])
@admin_required
def add_attendance_admin():
    user_id = request.form.get("user_id")
    data_str = request.form.get("data_sluzby")
    typ_mszy = request.form.get("typ_mszy")
    nazwa_inna = request.form.get("nazwa_inna")
    godzina = request.form.get("godzina")

    try:
        wybrana_data = date.fromisoformat(data_str)
        nowa_sluzba = Attendance(
            user_id=user_id, 
            data_sluzby=wybrana_data, 
            typ_mszy=typ_mszy,
            nazwa_inna=nazwa_inna if typ_mszy == 'inna' else None, 
            godzina=godzina
        )
        db.session.add(nowa_sluzba)
        db.session.commit()
        flash("Służba dodana pomyślnie przez Administratora.", "success")
    except Exception:
        db.session.rollback()
        flash("Coś poszło nie tak przy dodawaniu służby.", "danger")

    return redirect(url_for('admin_page'))

@app.route('/admin/add_announcement', methods=['POST'])
@staff_required
def add_announcement():
    tytul_val = request.form.get('tytul') or "Ogłoszenie"
    nowe = Announcement(tytul=tytul_val.strip(), tresc=request.form.get('tresc', '').strip())
    db.session.add(nowe)
    db.session.commit()
    flash("Nowe ogłoszenie dodane pomyślnie.", "success")
    if session.get('user_role') == 'ksiądz':
        return redirect(url_for('ksDash'))
    return redirect(url_for('admin_page'))

@app.route('/admin/edit_announcement/<int:id>', methods=['POST'])
@admin_required
def edit_announcement(id):
    ogloszenie = Announcement.query.get_or_404(id)
    ogloszenie.tytul = request.form.get('tytul', '').strip() or ogloszenie.tytul
    ogloszenie.tresc = request.form.get('tresc', '').strip()
    db.session.commit()
    flash("Ogłoszenie zaktualizowane.", "success")
    return redirect(url_for('admin_page'))

@app.route('/admin/delete_announcement/<int:id>', methods=['POST'])
@admin_required
def delete_announcement(id):
    ogloszenie = Announcement.query.get_or_404(id)
    db.session.delete(ogloszenie)
    db.session.commit()
    flash("Ogłoszenie usunięte.", "warning")
    return redirect(url_for('admin_page'))

@app.route('/admin/add_schedule', methods=['POST'])
@staff_required
def add_schedule():
    user_id = request.form.get('user_id')
    dzien = request.form.get('dzien')
    godzina = request.form.get('godzina')
    
    if not user_id or not dzien or not godzina:
        flash("Nie udało się zapisać. Wypełnij wszystkie pola formularza!", "warning")
        return redirect(url_for('admin_page'))
        
    try:
        nowy_dyzur = Schedule(
            user_id=int(user_id),
            dzien_tygodnia=dzien,
            godzina=godzina
        )
        db.session.add(nowy_dyzur)
        db.session.commit()
        flash("Służba dodana pomyślnie przez Administratora.", "success")
    except Exception:
        db.session.rollback()
        flash("Coś poszło nie tak przy dodawaniu służby.", "danger")
        
    return redirect(url_for('admin_page'))

@app.route('/admin/delete_schedule/<int:id>', methods=['POST'])
@staff_required
def delete_schedule(id):
    dyzur = db.session.get(Schedule, id)
    if dyzur:
        try:
            db.session.delete(dyzur)
            db.session.commit()
            flash("Służba została usunięta z grafiku.", "success")
        except Exception:
            db.session.rollback()
            flash("Błąd podczas usuwania służby.", "danger")
            
    return redirect(url_for('admin_page'))

@app.route('/admin')
@admin_required
def admin_page():
    all_attendance = db.session.query(Attendance, Users).join(Users).order_by(
        Attendance.data_sluzby.desc(), Attendance.godzina.desc()
    ).all()
    
    all_users = Users.query.filter_by(is_approved=True).all()
    pending_users = Users.query.filter_by(is_approved=False).order_by(Users.created_at.desc()).all()
    all_announcements = Announcement.query.order_by(Announcement.data_dodania.desc()).all()
    schedules = db.session.query(Schedule, Users).join(Users, Schedule.user_id == Users.id).all()

    wszystkie_dyzury = Schedule.query.all()
    aktualny_uzytkownik = db.session.get(Users, session['user_id'])
    
    plan_tygodnia = {'Poniedziałek': {}, 'Wtorek': {}, 'Środa': {}, 'Czwartek': {}, 'Piątek': {}, 'Sobota': {}, 'Niedziela': {}}
    plan_liczniki = {}
    for sch, u in schedules:
        dzien = sch.dzien_tygodnia
        godzina = sch.godzina
        if godzina not in plan_tygodnia[dzien]: plan_tygodnia[dzien][godzina] = []
        plan_tygodnia[dzien][godzina].append({'id': sch.id, 'user': u})
    for dzien in plan_tygodnia:
        plan_tygodnia[dzien] = dict(sorted(plan_tygodnia[dzien].items()))
        plan_liczniki[dzien] = sum(len(servers) for servers in plan_tygodnia[dzien].values())

    all_atts = Attendance.query.all()
    atts_by_user = defaultdict(list)
    for att in all_atts:
        atts_by_user[att.user_id].append(att)

    user_stats = []
    for u in all_users:
        his_atts = atts_by_user[u.id]
        total = len(his_atts)
        morning = sum(1 for a in his_atts if a.typ_mszy == 'poranna')
        evening = sum(1 for a in his_atts if a.typ_mszy == 'wieczorna')
        user_stats.append({
            'username': u.username, 'imie': u.imie, 'nazwisko': u.nazwisko,
            'full_name': f"{u.imie} {u.nazwisko}", 'uproszczony': u.uproszczony,
            'total': total, 'morning': morning, 'evening': evening, 'other': total - (morning + evening)
        })

    unapproved_users = Users.query.filter_by(is_approved=False).all()
    unapproved_count = len(unapproved_users)
    all_users = Users.query.order_by(Users.created_at.desc()).all()
    
    return render_template(
        "admin.html", 
        attendances=all_attendance, 
        users=all_users, 
        pending_users=pending_users,
        announcements=all_announcements, 
        stats=user_stats, 
        plan=plan_tygodnia, 
        liczniki=plan_liczniki,
        unapproved_users=unapproved_users,
        unapproved_count=unapproved_count,
        dyzury=wszystkie_dyzury,
        dyzury_count=len(wszystkie_dyzury),
        user=aktualny_uzytkownik.username
    )

@app.route('/admin/change_password_inline/<int:user_id>', methods=['POST'])
@admin_required
def admin_change_password_inline(user_id):
    user = Users.query.get_or_404(user_id)
    nowe_haslo = request.form.get('new_password')
    
    if nowe_haslo and nowe_haslo.strip() != "":
        user.password = nowe_haslo.strip()
        db.session.commit()
        flash(f'Hasło użytkownika {user.imie} {user.nazwisko} zostało zaktualizowane! 🔑', 'success')
    else:
        flash('Hasło nie może być puste!', 'danger')
        
    return redirect(url_for('admin_page'))

@app.route('/admin/approve_user/<int:user_id>', methods=['POST'])
@admin_required
def approve_user(user_id):
    user = Users.query.get_or_404(user_id)
    user.is_approved = True
    db.session.commit()
    flash(f'Konto użytkownika {user.username} zostało zatwierdzone!', 'success')
    return redirect(url_for('admin_page'))

@app.route('/admin/reject_user/<int:user_id>', methods=['POST'])
@admin_required
def reject_user(user_id):
    user = Users.query.get_or_404(user_id)
    Attendance.query.filter_by(user_id=user_id).delete()
    Schedule.query.filter_by(user_id=user_id).delete()
    db.session.delete(user)
    db.session.commit()
    flash(f'Konto użytkownika {user.username} zostało odrzucone i usunięte.', 'warning')
    return redirect(url_for('admin_page'))

@app.route('/admin/verify_action/<int:user_id>/<string:akcja>', methods=['POST'])
@admin_required
def verify_action(user_id, akcja):
    u = Users.query.get_or_404(user_id)
    if akcja == 'zatwierdz':
        u.is_approved = True
        flash(f"Użytkownik {u.username} zatwierdzony!", "success")
    else:
        Attendance.query.filter_by(user_id=user_id).delete()
        Schedule.query.filter_by(user_id=user_id).delete()
        db.session.delete(u)
        flash(f"Odrzucono rejestrację {u.username}.", "danger")
    db.session.commit()
    return redirect(url_for('admin_page'))

@app.route('/admin/delete_bulk_users', methods=['POST'])
@admin_required
def delete_bulk_users():
    user_ids = request.form.getlist('user_ids')
    if user_ids:
        for uid in user_ids:
            user_to_del = Users.query.get(uid)
            if user_to_del and user_to_del.role != 'admin':
                Attendance.query.filter_by(user_id=uid).delete()
                Schedule.query.filter_by(user_id=uid).delete()
                db.session.delete(user_to_del)
        db.session.commit()
        flash("Wybrani użytkownicy zostali trwale usunięci.", "success")
    else:
        flash("Proszę zaznaczyć użytkowników do usunięcia.", "warning")
    return redirect(url_for('admin_page'))

@app.route('/admin/delete_bulk_attendances', methods=['POST'])
@admin_required
def delete_bulk_attendances():
    att_ids = request.form.getlist('att_ids')
    if att_ids:
        for aid in att_ids:
            entry = Attendance.query.get(aid)
            if entry:
                db.session.delete(entry)
        db.session.commit()
        flash("Wybrane wpisy o służbach zostały usunięte.", "success")
    else:
        flash("Proszę zaznaczyć służby do usunięcia.", "warning")
    return redirect(url_for('admin_page'))

@app.route('/ksDash')
@staff_required
def ksDash():
    all_attendance = db.session.query(Attendance, Users).join(Users).order_by(
        Attendance.data_sluzby.desc(), Attendance.godzina.desc()
    ).all()
    all_users = Users.query.all()
    all_announcements = Announcement.query.order_by(Announcement.data_dodania.desc()).all()
    schedules = db.session.query(Schedule, Users).join(Users, Schedule.user_id == Users.id).all()
    
    plan_tygodnia = {'Poniedziałek': {}, 'Wtorek': {}, 'Środa': {}, 'Czwartek': {}, 'Piątek': {}, 'Sobota': {}, 'Niedziela': {}}
    plan_liczniki = {}

    for sch, u in schedules:
        dzien = sch.dzien_tygodnia
        godzina = sch.godzina
        if godzina not in plan_tygodnia[dzien]:
            plan_tygodnia[dzien][godzina] = []
        plan_tygodnia[dzien][godzina].append({'id': sch.id, 'user': u})
        
    for dzien in plan_tygodnia:
        plan_tygodnia[dzien] = dict(sorted(plan_tygodnia[dzien].items()))
        count = sum(len(servers) for servers in plan_tygodnia[dzien].values())
        plan_liczniki[dzien] = count
    
    user_atts_map = {u.id: [] for u in all_users}
    for att, usr in all_attendance:
        if usr.id in user_atts_map:
            user_atts_map[usr.id].append(att)

    user_stats = []
    for u in all_users:
        his_atts = user_atts_map.get(u.id, [])
        total = len(his_atts)
        morning = sum(1 for a in his_atts if a.typ_mszy == 'poranna')
        evening = sum(1 for a in his_atts if a.typ_mszy == 'wieczorna')
        other = total - (morning + evening)
        
        user_stats.append({
            'username': u.username, 
            'imie': u.imie,
            'nazwisko': u.nazwisko,
            'full_name': f"{u.imie} {u.nazwisko}",
            'total': total, 
            'morning': morning, 
            'evening': evening, 
            'other': other
        })

    return render_template(
        'ks.html', 
        attendances=all_attendance, 
        users=all_users, 
        announcements=all_announcements, 
        stats=user_stats, 
        plan=plan_tygodnia, 
        liczniki=plan_liczniki
    )

@app.route('/dashboard', methods=['GET', 'POST'])
@login_required
def dashboard_page():
    user_id = session['user_id']
    aktualny_uzytkownik = db.session.get(Users, user_id)
    if not aktualny_uzytkownik:
        return redirect(url_for('logout'))

    if request.method == 'POST':
        data_sluzby_str = request.form.get('data_sluzby')
        godzina = request.form.get('godzina')
        typ_mszy = request.form.get('typ')
        nazwa_inna = request.form.get('inna') if typ_mszy == 'inna' else None
        
        try:
            data_sluzby = datetime.strptime(data_sluzby_str, '%Y-%m-%d').date()
            dzisiaj = date.today()
            wczoraj = dzisiaj - timedelta(days=1)
            hard_limit = date(2026, 4, 12)

            if data_sluzby > dzisiaj or data_sluzby < max(wczoraj, hard_limit):
                flash("Wybrana data służby jest nieprawidłowa.", "danger")
                return redirect(url_for('dashboard_page'))

            istniejaca = Attendance.query.filter_by(
                user_id=user_id, 
                data_sluzby=data_sluzby, 
                godzina=godzina
            ).first()

            if istniejaca:
                flash("Masz już zgłoszoną służbę o tej godzinie.", "danger")
                return redirect(url_for('dashboard_page'))

            new_attendance = Attendance(
                user_id=user_id,
                data_sluzby=data_sluzby,
                typ_mszy=typ_mszy,
                nazwa_inna=nazwa_inna,
                godzina=godzina
            )
            db.session.add(new_attendance)
            db.session.commit()
            flash('Obecność na służbie zapisana pomyślnie.', 'success')
        except Exception:
            db.session.rollback()
            flash("Coś poszło nie tak przy zapisywaniu służby.", "danger")
        return redirect(url_for('dashboard_page'))
        
    announcements = Announcement.query.order_by(Announcement.id.desc()).all()
    attendances = Attendance.query.filter_by(user_id=user_id).order_by(Attendance.data_sluzby.desc()).all()
    
    moje_dyzury = Schedule.query.filter_by(user_id=user_id).order_by(Schedule.dzien_tygodnia, Schedule.godzina).all()
    schedules = Schedule.query.options(joinedload(Schedule.user)).all()
    
    dni_kolejnosc = ['Poniedziałek', 'Wtorek', 'Środa', 'Czwartek', 'Piątek', 'Sobota', 'Niedziela']
    plan_tygodnia = {dzien: {} for dzien in dni_kolejnosc}
    plan_liczniki = {dzien: 0 for dzien in dni_kolejnosc}
    
    for sch in schedules:
        dzien = sch.dzien_tygodnia
        godzina = sch.godzina
        if dzien in plan_tygodnia:
            if godzina not in plan_tygodnia[dzien]:
                plan_tygodnia[dzien][godzina] = []
            plan_tygodnia[dzien][godzina].append(sch.user)
            plan_liczniki[dzien] += 1
            
    for dzien in plan_tygodnia:
        plan_tygodnia[dzien] = dict(sorted(plan_tygodnia[dzien].items()))
    
    if aktualny_uzytkownik.uproszczony:
        return render_template('dash_uproszczony.html', user=aktualny_uzytkownik.username, announcements=announcements, attendances=attendances)
    return render_template('dashboard.html', user=aktualny_uzytkownik.username, announcements=announcements, attendances=attendances, moje_dyzury=moje_dyzury, plan=plan_tygodnia, liczniki=plan_liczniki)

@app.route('/delete_my_attendance/<int:id>', methods=['POST'])
@login_required
def delete_my_attendance(id):
    entry = Attendance.query.get_or_404(id)
    if entry.user_id != session['user_id']:
        flash("Brak uprawnień. Nie można usunąć cudzego wpisu.", "danger")
        return redirect(url_for('dashboard_page'))
    try:
        db.session.delete(entry)
        db.session.commit()
        flash("Wpis o Twojej służbie został usunięty.", "success")
    except Exception:
        db.session.rollback()
        flash("Coś poszło nie tak przy usuwaniu służby.", "danger")
    return redirect(url_for('dashboard_page'))

@app.route('/edit_my_attendance/<int:id>', methods=['POST'])
@login_required
def edit_my_attendance(id):
    att = Attendance.query.get_or_404(id)
    if att.user_id != session['user_id']:
        flash('Brak uprawnień do edycji tej służby.', 'danger')
        return redirect(url_for('dashboard_page'))
        
    try:
        att.data_sluzby = datetime.strptime(request.form.get('data_sluzby'), '%Y-%m-%d').date()
        att.godzina = request.form.get('godzina')
        att.typ_mszy = request.form.get('typ')
        att.nazwa_inna = request.form.get('inna') if att.typ_mszy == 'inna' else None
        
        db.session.commit()
        flash('Zgłoszenie zostało pomyślnie zaktualizowane.', 'success')
    except Exception:
        db.session.rollback()
        flash('Coś poszło nie tak przy edycji służby.', 'danger')
        
    return redirect(url_for('dashboard_page'))

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login_page'))

@app.route('/forget-password')
def forget_password():
    return render_template('forget-password.html')

@app.route('/robots.txt')
def static_from_root():
    return send_from_directory(app.static_folder, 'robots.txt')

@app.route('/sitemap.xml')
def sitemap_from_root():
    return send_from_directory(app.static_folder, 'sitemap.xml')

@app.route('/admin/export_attendances')
@staff_required
def export_attendances():
    attendances = db.session.query(Attendance, Users).join(Users).order_by(
        Attendance.data_sluzby.desc(), Attendance.godzina.desc()
    ).all()

    wb = openpyxl.Workbook()
    
    ws = wb.active
    ws.title = "Lista Służb"
    ws.views.sheetView[0].showGridLines = True

    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="5E3BEE", end_color="5E3BEE", fill_type="solid")
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB')
    )

    headers = ["ID", "Imię i Nazwisko", "Login", "Data Służby", "Godzina", "Typ Liturgii", "Nazwa Inna / Uwagi"]
    ws.append(headers)

    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = thin_border

    typ_counts = defaultdict(int)

    for idx, (att, u) in enumerate(attendances, start=2):
        typ_label = att.typ_mszy.capitalize() if att.typ_mszy else "Nieokreślony"
        typ_counts[typ_label] += 1

        row = [
            att.id,
            f"{u.imie} {u.nazwisko}",
            u.username,
            att.data_sluzby.strftime("%Y-%m-%d") if att.data_sluzby else "",
            att.godzina,
            typ_label,
            att.nazwa_inna or ""
        ]
        ws.append(row)

        for col_num in range(1, len(headers) + 1):
            c = ws.cell(row=idx, column=col_num)
            c.border = thin_border
            c.alignment = align_center if col_num in [1, 4, 5, 6] else align_left

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    ws_chart = wb.create_sheet(title="Statystyki i Wykres")
    ws_chart.views.sheetView[0].showGridLines = True

    ws_chart.append(["Typ Liturgii", "Liczba Służb"])
    ws_chart.cell(row=1, column=1).font = header_font
    ws_chart.cell(row=1, column=1).fill = header_fill
    ws_chart.cell(row=1, column=2).font = header_font
    ws_chart.cell(row=1, column=2).fill = header_fill

    r_idx = 2
    for t_name, count in typ_counts.items():
        ws_chart.append([t_name, count])
        ws_chart.cell(row=r_idx, column=1).border = thin_border
        ws_chart.cell(row=r_idx, column=2).border = thin_border
        r_idx += 1

    pie = PieChart()
    pie.title = "Podział Służb ze względu na Typ"
    labels = Reference(ws_chart, min_col=1, min_row=2, max_row=r_idx - 1)
    data = Reference(ws_chart, min_col=2, min_row=1, max_row=r_idx - 1)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    ws_chart.add_chart(pie, "D4")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="sluzby_export.xlsx"
    )

@app.route("/admin/export_ranking")
@staff_required
def export_ranking():
    users = Users.query.filter(Users.role == 'user').all()
    attendances = Attendance.query.all()
    
    morning_count = defaultdict(int)
    evening_count = defaultdict(int)
    other_count = defaultdict(int)
    total_count = defaultdict(int)
    
    for att in attendances:
        total_count[att.user_id] += 1
        time_str = att.godzina.replace(":", "")
        try:
            hour_val = int(time_str[:2])
            if hour_val < 12:
                morning_count[att.user_id] += 1
            elif hour_val >= 17:
                evening_count[att.user_id] += 1
            else:
                other_count[att.user_id] += 1
        except ValueError:
            other_count[att.user_id] += 1

    data = []
    for u in users:
        tot = total_count[u.id]
        level = "LIDER" if tot >= 20 else ("AKTYWNY" if tot >= 5 else "POCZĄTKUJĄCY")
        data.append({
            "Login": u.username,
            "Imię i Nazwisko": f"{u.imie} {u.nazwisko}",
            "Suma Służb (Punkty)": tot,
            "Służby Poranne": morning_count[u.id],
            "Służby Wieczorne": evening_count[u.id],
            "Inne / Dodatkowe": other_count[u.id],
            "Status Aktywności": level
        })
        
    df = pd.DataFrame(data).sort_values(by="Suma Służb (Punkty)", ascending=False)
    buffer = io.BytesIO()
    df.to_excel(buffer, index=False, sheet_name="Cyfrowy Ranking")
    buffer.seek(0)
    
    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"cyfrowy_ranking_{date.today()}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.route('/export_schedule')
@login_required
def export_schedule():
    schedules = db.session.query(Schedule, Users).join(Users, Schedule.user_id == Users.id).all()
    
    data = []
    dni_kolejnosc = {'Poniedziałek': 1, 'Wtorek': 2, 'Środa': 3, 'Czwartek': 4, 'Piątek': 5, 'Sobota': 6, 'Niedziela': 7}
    
    for sch, u in schedules:
        data.append({
            'Dzień Tygodnia': sch.dzien_tygodnia,
            'Kolejność': dni_kolejnosc.get(sch.dzien_tygodnia, 8),
            'Godzina': sch.godzina,
            'Imię i Nazwisko': f"{u.imie} {u.nazwisko}",
            'Pseudonim (Login)': u.username
        })

    df = pd.DataFrame(data)
    if not df.empty:
        df = df.sort_values(by=['Kolejność', 'Godzina'])
        df = df.drop(columns=['Kolejność'])
    else:
        df = pd.DataFrame(columns=['Dzień Tygodnia', 'Godzina', 'Imię i Nazwisko', 'Pseudonim (Login)'])

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Plan_Sluzb')
    
    output.seek(0)
    nazwa_pliku = f"Plan_Sluzb_{date.today().strftime('%Y-%m-%d')}.xlsx"
    return send_file(output, download_name=nazwa_pliku, as_attachment=True)

@app.route('/admin/toggle_uproszczony/<int:id>', methods=['POST'])
@admin_required
def toggle_uproszczony(id):
    u = Users.query.get_or_404(id)
    u.uproszczony = not u.uproszczony
    db.session.commit()
    flash(f"Zmieniono tryb wyświetlania dla użytkownika {u.username}.", "success")
    return redirect(url_for('admin_page'))

# --- ZABEZPIECZENIE SERWOWANIA PLIKU PRZED PATH TRAVERSAL ---
@app.route('/download/regulamin.pdf')
def pobierz_regulamin():
    katalog = os.path.join(app.root_path, 'static', 'docs')
    safe_path = os.path.abspath(os.path.join(katalog, 'regulamin.pdf'))
    if not safe_path.startswith(os.path.abspath(katalog)):
        abort(403)
    return send_from_directory(
        katalog, 
        'regulamin.pdf', 
        as_attachment=True, 
        download_name='regulamin.pdf'
    )

@app.route('/pomoc')
def pomoc_page():
    return render_template('pomoc.html')

with app.app_context(): 
    db.create_all()
    env_admin_name = os.getenv("ADMIN_NAME") or os.getenv("admin_name") or "AdminGreg"
    env_admin_pass = os.getenv("ADMIN_PASSWORD") or os.getenv("admin_password") or "Lego2012"
    
    admin_user = Users.query.filter_by(username=env_admin_name).first()
    if not admin_user:
        hashed_admin_pass = generate_password_hash(env_admin_pass, method='pbkdf2:sha256')
        admin_user = Users(
            imie="Główny", 
            nazwisko="Szef", 
            username=env_admin_name, 
            password=hashed_admin_pass, 
            role='admin', 
            uproszczony=False,
            is_approved=True
        )
        db.session.add(admin_user)
        db.session.commit()
    elif not admin_user.password.startswith('pbkdf2:'):
        admin_user.password = generate_password_hash(env_admin_pass, method='pbkdf2:sha256')
        admin_user.is_approved = True
        db.session.commit()

if __name__ == '__main__':
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)